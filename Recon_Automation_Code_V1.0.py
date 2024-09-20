from openpyxl import load_workbook
import os
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import win32com.client as client
import shutil
import re
import subprocess
import zipfile

#///////////////////////////////////////////////////////////////////////
#  Unzip the file then UnRar the file 

def extract_archive(input_folder, output_folder):
    if input_folder.endswith(".zip"):
        shutil.unpack_archive(input_folder, output_folder, format='zip')
    else:
        print(f"Unsupported file format: {input_folder}")

# Replace 'your_archive_file.zip' with the actual file path
#input_folder = "[EXTERNAL]_Reconciliation_with_Waybill_January_24,_2024.zip"
# Get the folder path from user input
#path = input("Enter File Path: ").strip()    "C:\Users\Shah\Downloads\jan 24 kpi 2223"
File_Name = input("Enter file Name: \n")
# Extract the file name from the file path
file_name = os.path.basename(File_Name)
    
# Extract the month name, date, and year from the file name using regular expressions
match = re.search(r'([a-zA-Z]+)_([0-9]+),_([0-9]+)', file_name)
if match:
    month_name = match.group(1)
    date = match.group(2)
    year = match.group(3)
    print(f"Month Name: {month_name}")
    print(f"Date: {date}")
    print(f"Year: {year}")
else:
    print("Date, month, and year not found in file name")

input_folder = os.path.join(os.getcwd(), f"C:\\Users\\Shah\\Downloads\\{File_Name}")
#input_folder = os.path.join(os.getcwd(), path)
output_folder = "Unzip_output_folder"

if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Extract the main archive
extract_archive(input_folder, output_folder)
# List the contents of the output folder after the first extraction
print(f"Contents of {output_folder}: {os.listdir(output_folder)} \n")

#/////////////////////////////////////////////////////////////////////////////////////////////////////////////
def extract_rar_with_winrar(winrar_path, input_folder, output_folder):
    # Check if the output folder exists, if not, create it
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Check if the 'UNRAR New' directory exists, if not, create it
    unrar_new_folder = os.path.join(output_folder, "RE UNRAR NEW")
    if not os.path.exists(unrar_new_folder):
        os.makedirs(unrar_new_folder)

    # Verify if WinRAR.exe exists at the specified path
    if not os.path.isfile(winrar_path):
        print(f"Error: WinRAR.exe not found at the specified path: {winrar_path}")
        return

    for file in os.listdir(input_folder):
        if file.endswith(".rar"):
            rar_path = os.path.join(input_folder, file)
            # Run the WinRAR command to extract the RAR file
            subprocess.run([winrar_path, "x", rar_path, unrar_new_folder])

    # Iterate over the files in the "UNRAR New" folder for further extraction
    for nested_file in os.listdir(unrar_new_folder):
        nested_rar_path = os.path.join(unrar_new_folder, nested_file)
        if nested_file.endswith(".rar"):
            # Run the WinRAR command to extract the nested RAR file
            subprocess.run([winrar_path, "x", nested_rar_path, output_folder])

# Replace 'your_winrar_path' and 'Unzip_output_folder' with the actual paths
winrar_path = "C:\\Program Files\\WinRAR\\WinRAR.exe"
input_folder = "Unzip_output_folder"
output_folder = "RE UNRAR NEW"

extract_rar_with_winrar(winrar_path, input_folder, output_folder)

shutil.rmtree('Unzip_output_folder')


input_folder = "RE UNRAR NEW"
output_folder = "Converted Only XLS files"

# Create the output folder if it doesn't exist
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Get a list of all files in the input folder
all_files = os.listdir(input_folder)

# Filter only .xls files
xls_files = [file for file in all_files if file.lower().endswith('.xls')]

# Move each .xls file to the output folder
for xls_file in xls_files:
    source_path = os.path.join(input_folder, xls_file)
    destination_path = os.path.join(output_folder, xls_file)
    shutil.move(source_path, destination_path)

print(f"Moved {len(xls_files)} .xls files to '{output_folder}' folder.")

shutil.rmtree('RE UNRAR NEW')

#///////////////////////////////////////////////////////////////////////



# To Change the extension of Excel From >> .XLS TO >> .XLSX

excel = client.Dispatch("excel.application")

folder_path = os.path.join(os.getcwd(), "Converted Only XLS files")
output_folder = os.path.join(os.getcwd(), "CONVERTED Waybills Chamkani xlsx")

if not os.path.exists(output_folder):
    os.makedirs(output_folder)

for file in os.listdir(folder_path):
    filename, file_extension = os.path.splitext(file)
    file_path = os.path.join(folder_path, file)

    try:
        # Open the workbook
        wb = excel.workbooks.open(file_path)
        
        if wb:
            # Save it as xlsx
            output_path = os.path.join(output_folder, f"{filename}.xlsx")
            wb.SaveAs(output_path, FileFormat=51)  # 51 represents xlsx format
            wb.Close()
        else:
            print(f"Error processing {file}: Unable to open workbook.")
    except Exception as e:
        print(f"Error processing {file}: {e}")

excel.Quit()

shutil.rmtree('Converted Only XLS files')
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

                                                                ##################################
                                                                ##################################
                                                                #     FUNCTION DEFINE SPACE      #
                                                                ##################################
                                                                ##################################

#This Fucntion will take the File Path argument and then it will apply all the First step code to all the Excel Files present in the folders. 
def process_excel_file(file_path, output_folder):
    # Load the Excel workbook
    book = load_workbook(file_path)

    # Perform operations on the workbook, for example:
    sheet = book.active
    book = load_workbook(file_path)
    sheet = book.active
    sheet_names = book.sheetnames
    # Perform operations on the workbook, for example:
    filename = file_path.split('/')[-1]  # Assuming the file path uses '/' as separator
    print("Filename:", filename)
   
    if 'UpLink' in sheet_names and 'Downlink' in sheet_names:

        # Get references to the sheets
        downlink_sheet = book['Downlink']
        uplink_sheet = book['UpLink']

        # Iterate through the rows in the Downlink sheet and append to the Uplink sheet
        for row in downlink_sheet.iter_rows(min_row=2, values_only=True):
            uplink_sheet.append(row)

        if 'Downlink' in sheet_names:
            book.remove(downlink_sheet)  # Remove the Downlink Sheet
            print("Downlink Successfully Removed!")
            print("Data is successfully added to Uplink!")
# Code for removing Duplicate Trips/rows if any
#////////////////////////////
            # Create a set to store unique rows
            unique_rows = set()

            # List to store rows to be removed
            rows_to_remove = []

            # Iterate through rows in "UpLink" sheet
            for row in uplink_sheet.iter_rows(min_row=2, values_only=True):
                if row not in unique_rows:
                    # Add the row to the set of unique rows
                    unique_rows.add(row)
                else:
                    # Add the row index to rows_to_remove if it's a duplicate
                    rows_to_remove.append(row)

            # Remove duplicate rows from the "UpLink" sheet
            for row in rows_to_remove:
                for r in uplink_sheet.rows:
                    if r[0].value == row[0] and r[1].value == row[1]:
                        uplink_sheet.delete_rows(r[0].row)  
        #///////////////////////////   
            # Delete specified columns
            sheet = book.active
            sheet.delete_cols(3)
            sheet.delete_cols(3)
            sheet.delete_cols(4)
            sheet.delete_cols(6)
            sheet.insert_cols(8)  # Insert a column at index 8 (between G and H)
            sheet.delete_cols(13)
            sheet.delete_cols(13)
            sheet.delete_cols(13)
            sheet.delete_cols(13)
        
            print("Columns are successfully Removed!")
            # Insert a new row at the top
            sheet.insert_rows(1)
            print("Row on the top is successfully inserted!")
            #///////////////////////////////////////////////////////////////
            #Entering 'Verified' and 'Waybill' to a cell of Status and Verified From
            last_row = uplink_sheet.max_row

            # Set the value "Verified" in cell O3
            uplink_sheet['O3'] = 'Verified'
            uplink_sheet['P3'] = 'Waybill'

            # Fill down the value in column O from O3 to the last row
            for row in uplink_sheet.iter_rows(min_row=3, max_row=last_row, min_col=15, max_col=15):
                for cell in row:
                    cell.value = 'Verified'
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Fill down the value in column P from P3 to the last row
            for row in uplink_sheet.iter_rows(min_row=3, max_row=last_row, min_col=16, max_col=16):
                for cell in row:
                    cell.value = 'Waybill'
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            print("'Verified' and 'Waybill' are successfully inserted till the last cell, and filter is applied!")
            
            #///////////////////////////////////////////////////////////////////////

            #Adding boarder From U2 to Y12
            for row in sheet.iter_rows(min_row=2, max_row=12, min_col=21, max_col=25):
                for cell in row:
                 border_style = Side(style='thin')
                 cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            #Adding Boarder to from U2 to Y12& one cell of Y13
            print("table is addes successfully")
            
            #Adding Data to the New Table 13
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            # Apply the border and fill to Y13 cell
            sheet.cell(row=13, column=25).border = thin_border
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            y13_cell = sheet.cell(row=13, column=25)
            y13_cell.border = thin_border
            y13_cell.fill = yellow_fill

            #Filling color to Table 
            #///////////////////////////////////////////////////////
            # Specify the range of cells (U2 to Y12)
            start_cell = sheet['U2']
            end_cell = sheet['Y12']
            target_range = sheet[start_cell.coordinate:end_cell.coordinate]
            # Create a PatternFill object with a green color
            fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            # Apply the PatternFill object to the range of cells
            for row in target_range:
                for cell in row:
                    cell.fill = fill
                    
            
            #////////////////////////////////////////////////////////
            # Table Data Manopulation Adding Data to Table
            sheet.merge_cells('U2:V2') #this will merg the cells 
            c7 =sheet.cell(row=2, column=21)
            c7.value = "12m"
            #////////////////////////////////////////////////////
            # This will Center the Vlaue in the specify Cell
            target_cell = sheet['U2']
            alignment = Alignment(horizontal='center', vertical='center')
            target_cell.alignment = alignment 
        #////////////////////////////////////////////////////    
            c7 =sheet.cell(row=3, column=21)
            c7.value = "FWD"
            c7 =sheet.cell(row=6, column=21)
            c7.value = "RDs"
            c7 =sheet.cell(row=7, column=21)
            c7.value = "12M"
            c7 =sheet.cell(row=10, column=21)
            c7.value = "12M"
            c7 =sheet.cell(row=3, column=22)
            c7.value = "BWD"
            sheet.merge_cells('W2:X2') #this will merg the cells 
            c8 =sheet.cell(row=2, column=23)
            c8.value = "18m"
            #////////////////////////////////////////////////////
            # This will Center the Vlaue in the specify Cell
            target_cell = sheet['W2']
            alignment = Alignment(horizontal='center', vertical='center')
            target_cell.alignment = alignment
            #////////////////////////////////////////////////////////   
            c8 =sheet.cell(row=3, column=23)
            print("Data to table is addes successfully")
            c8.value = "FWD"
            c8 =sheet.cell(row=10, column=23)
            c8.value = "18M"
            c8 =sheet.cell(row=3, column=24)
            c8.value = "BWD"
            c8 =sheet.cell(row=7, column=24)
            c8.value = "18M"
            c9 =sheet.cell(row=3, column=25)
            c9.value = "Total"
            c9 =sheet.cell(row=8, column=25)
            c9.value = "Total"
            c9 =sheet.cell(row=11, column=25)
            c9.value = "Total"

            #///////////////////////////////////////////////////////
            #Adding data to First row of sheet and Formating
            sheet.merge_cells('O1:P1') #this will merg the cells 
            sheet.merge_cells('R1:S1')
            c1 =sheet.cell(row=1, column=15)
            c1.value = "LMKR"
            c2 =sheet.cell(row=1, column=18)
            c2.value = "KPIs"
            c3 =sheet.cell(row=1, column=20)
            c3.value = "TPC" 
            c4 =sheet.cell(row=2, column=15)
            c4.value = "Status"
            c5 =sheet.cell(row=2, column=16)
            c5.value = "Verified From"
            c6 =sheet.cell(row=2, column=18)
            c6.value = "Headway"
            c6 =sheet.cell(row=2, column=19)
            c6.value = "Travel Time"
            c6 =sheet.cell(row=2, column=8)
            c6.value = "Travel Time"
            #Making Cell Values BOLD
            sheet['A1'].font = Font(b=True)
            sheet['T1'].font = Font(b=True)
            sheet['R1'].font = Font(b=True)
            sheet['O1'].font = Font(b=True)
            #////////////////////////////////////////////////////////////////
            
            last_row = sheet.max_row
            # Apply the formula to all cells in column H from H3 to the last filled row
            for row in range(3, last_row + 1):
                sheet[f'H{row}'] = f'=G{row}-F{row}'
            
            # Apply the time format to column H
            time_style = NamedStyle(name='time_style', number_format='[hh]:mm')
            for row in range(3, last_row + 1):
                sheet[f'H{row}'].style = time_style
            
          #////////////////////////////////////////////////////////////////////
            print("Data is successfully inserted to the Cells! \n \n")

    # Save the changes in the new folder
    output_file_path = os.path.join(output_folder, os.path.basename(file_path))
    
    book.save(output_file_path)

    # Close the workbook
    book.close()

#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#This Function will read all the files in the folder then will filter out the excel files from it 
#Then it will store the Excel file in a file_path variable that we will be passing to a new functin called process_excel_file() 
def process_excel_files_in_folder(input_folder, output_folder):
    # List all files in the input folder
    files = [f for f in os.listdir(input_folder) if os.path.isfile(os.path.join(input_folder, f))]

    # Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]

    # Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(input_folder, excel_file)
        print(f"Processing {excel_file}...")
        process_excel_file(file_path, output_folder)

#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#These functions are only For Chamkani Routes 


# Define functions for each file
                                  #      /////////////////////////////////  ROUTES FOR CHAMAKANI  ///////////////////////////////////////////////  

def dr3a():
    
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'DR-03A'
    dr03a_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'DR-04b'
        if 'DR-03A' in excel_file:
            dr03a_file = excel_file
            print(f"File with 'DR-03A' found: {dr03a_file}")
            
# Check if a file with 'DR-03A' was found
    if dr03a_file:
        try:
        # Read the Excel file and get the value of cell A2
            dr03a_file_path = os.path.join(output_folder_path, dr03a_file)
            workbook = load_workbook(dr03a_file_path)
            sheet = workbook.active
            sheet.title = 'DR-03A'
            #Extracting Date From File name using String indexing
            print("Filename:", dr03a_file)
            date = dr03a_file[25:33]
            print(f"Date: {date}")      # This will print Date from File name 
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
            
            #////////////////////////////////////////////////////////////////////
            # Apply the formula to all THE GREEN cells in column U4 to Y13
            # GREEN TABLE DATA FOR DR-03A 
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD DG"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD Kohat Adda"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD DG"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD Kohat Adda"
            sheet = workbook.active
            sheet.title = 'DR-03A'
            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD DG",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Kohat Adda",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Kohat Adda",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
        # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD DG', 'Backward RD DG', 'Forward RD Kohat Adda']
            L2 = ['FWD', 'BWD', 'FWD RD DG', 'BWD RD DG', 'FWD RD KOHAT ADDA']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)

#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD DG']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD DG']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD KOHAT ADDA']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
  
        except Exception as e:
            print(f"Error reading {dr03a_file}: {e}")
    else:
        print("No file with 'DR-03A' found. \n")
    workbook.save(dr03a_file_path)    
    print("DR-03A file Code Exicuted And \n Date is added And \n Sheets are added \n")     



def dr3b():
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'DR-04B'
    dr03b_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'DR-03B'
        if 'DR-03B' in excel_file:
            dr03b_file = excel_file
            print(f"File with 'DR-03B' found: {dr03b_file}")
            
# Check if a file with 'DR-03B' was found
    if dr03b_file:
        try:
        # Read the Excel file and get the value of cell A2
            dr03b_file_path = os.path.join(output_folder_path, dr03b_file)
            workbook = load_workbook(dr03b_file_path)
            sheet = workbook.active
            print("Filename:", dr03b_file)
            date = dr03b_file[25:33]
            print(f"Date: {date}")      # This will print Date from File name 
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell

            # GREEN TABLE DATA FOR DR-03A 
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD MSS"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD SHP"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD MSS"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD SHP"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'

            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD MSS",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD MSS",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD SHP",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            sheet = workbook.active
            sheet.title = 'DR-03B'
        # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD MSS', 'Backward RD MSS', 'Forward RD SHP', 'Backward RD SHP']
            L2 = ['FWD', 'BWD', 'FWD RD MSS', 'BWD RD MSS', 'FWD RD SHP', 'BWD RD SHP']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)

#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD MSS']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD MSS']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD SHP']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD SHP']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
		   
            

        except Exception as e:
            print(f"Error reading {dr03b_file}: {e}")
    else:
        print("No file with 'DR-03B' found. \n")
    workbook.save(dr03b_file_path)    
    print("DR-03B file Code Exicuted And \n Date is added And \n Sheets are added \n") 
    


def dr4b():
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'DR-04B'
    dr04b_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'DR-04b'
        if 'DR-04B' in excel_file:
            dr04b_file = excel_file
            print(f"File with 'DR-04B' found: {dr04b_file}")
            
# Check if a file with 'DR-04B' was found
    if dr04b_file:
        try:
        # Read the Excel file and get the value of cell A2
            dr04b_file_path = os.path.join(output_folder_path, dr04b_file)
            workbook = load_workbook(dr04b_file_path)
            sheet = workbook.active
            sheet.title = 'DR-04B'
            print("Filename:", dr04b_file)
            date = dr04b_file[25:33]
            print(f"Date: {date}")      # This will print Date from File name 
            sheet = workbook.active
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
            
            #/////////////////////////////////
		
            # GREEN TABLE DATA FOR DR-04B
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD MSS"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "N/A"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD MSS"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "N/A"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD MSS",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD MSS",$O:$O,"V*")'

            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'

            sheet = workbook.active
        # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward']
            L2 = ['FWD', 'BWD']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)

#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
  
        except Exception as e:
            print(f"Error reading {dr04b_file}: {e}")
    else:
        print("No file with 'DR-04B' found. \n")
        
    workbook.save(dr04b_file_path)    
    print("DR-04B file Code Exicuted And \n Date is added And \n Sheets are added \n") 

    
   
def er01():
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'ER-01'
    er01_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'ER-01'
        if 'ER-01' in excel_file:
            er01_file = excel_file
            print(f"File with 'ER-01' found: {er01_file}")
            
# Check if a file with 'DR-01' was found
    if er01_file:
        try:
        # Read the Excel file and get the value of cell A2
            er01_file_path = os.path.join(output_folder_path, er01_file)
            workbook = load_workbook(er01_file_path)
            sheet = workbook.active	
            sheet.title = 'ER-01'
            print("Filename:", er01_file)
            date = er01_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
            # GREEN TABLE DATA FOR ER-01
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "N/A"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "N/A"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "N/A"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "N/A"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD MSS",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD MSS",$O:$O,"V*")'

            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
        # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward']
            L2 = ['FWD', 'BWD']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

            #//////////////////////////////////////////////////////////////


        except Exception as e:
            print(f"Error reading {er01_file}: {e}")
    else:
        print("No file with 'ER-01' found. \n")
    workbook.save(er01_file_path)    
    print("DR-03A file Code Exicuted And \n Date is added And \n Sheets are added \n")     



def er10():
    
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'ER-10'
    er10_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'ER-10'
        if 'ER-10' in excel_file:
            er10_file = excel_file
            print(f"File with 'ER-10' found: {er10_file}")
            
# Check if a file with 'ER-10' was found
    if er10_file:
        try:
        # Read the Excel file and get the value of cell A2
            er10_file_path = os.path.join(output_folder_path, er10_file)
            workbook = load_workbook(er10_file_path)
            sheet = workbook.active
            sheet.title = 'ER10'
            print("Filename:", er10_file)
            date = er10_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 

		
            # GREEN TABLE DATA FOR ER-10
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD Kohat Adda"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD Hospital Chowk"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD Kohat Adda"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD Hospital Chowk"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Kohat Adda",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Kohat Adda",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD HC",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD HC",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD HC",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD HC",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
        # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD Kohat Adda', 'Backward RD Kohat Adda', 'Forward RD HC', 'Backward RD HC']
            L2 = ['FWD', 'BWD', 'FWD RD Kohat Adda', 'BWD RD Kohat Adda', 'Forward RD HC', 'Backward RD HC']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD Kohat Adda']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD Kohat Adda']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['Forward RD HC']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['Backward RD HC']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        except Exception as e:
            print(f"Error reading {er10_file}: {e}")
    else:
        print("No file with 'ER-10' found. \n")
    workbook.save(er10_file_path)    
    print("ER-10 file Code Exicuted And \n Date is added And \n Sheets are added \n")     



def er12():
    
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'ER-12'
    er12_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'ER-12'
        if 'ER-12' in excel_file:
            er12_file = excel_file
            print(f"File with 'ER-12' found: {er12_file}")
            
# Check if a file with 'DR-12' was found
    if er12_file:
        try:
        # Read the Excel file and get the value of cell A2
            er12_file_path = os.path.join(output_folder_path, er12_file)
            workbook = load_workbook(er12_file_path)
            sheet = workbook.active
            print("Filename:", er12_file)
            date = er12_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
            sheet.title = 'ER12'
		
            # GREEN TABLE DATA FOR ER-12
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD MOH"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD SHP"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD MOH"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD SHP"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD MOH",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD SHP",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'

        # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD MOH', 'Backward RD MOH', 'Forward RD SHP', 'Backward RD SHP']
            L2 = ['FWD', 'BWD', 'FWD RD MOH', 'BWD RD MOH', 'FWD RD SHP', 'BWD RD SHP']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD MOH']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD MOH']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD SHP']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD SHP']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

        except Exception as e:
            print(f"Error reading {er12_file}: {e}")
    else:
        print("No file with 'ER-12' found. \n")
    workbook.save(er12_file_path)    
    print("ER-12 file Code Exicuted And \n Date is added And \n Sheets are added \n")     



def xer15():
     
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'XER-15'
    Xer15_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'XER-15'
        if 'XER-15' in excel_file:
            Xer15_file = excel_file
            print(f"File with 'XER-15' found: {Xer15_file}")
            
# Check if a file with 'XER-15' was found
    if Xer15_file:
        try:
        # Read the Excel file and get the value of cell A2
            Xer15_file_path = os.path.join(output_folder_path, Xer15_file)
            workbook = load_workbook(Xer15_file_path)
            sheet = workbook.active
            sheet.title = 'XER15'
            print("Filename:", Xer15_file)
            date = Xer15_file[25:33]
            print(f"Date: {date}")      # This will print Date from File name 
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell

            # GREEN TABLE DATA FOR XER-15
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "N/A"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "N/A"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "N/A"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "N/A"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD DG",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD DG",$O:$O,"V*")'

            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD SHP",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward']
            L2 = ['FWD', 'BWD']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
  

        except Exception as e:
            print(f"Error reading {Xer15_file}: {e}")
    else:
        print("No file with 'XER-15' found. \n")
    workbook.save(Xer15_file_path)    
    print("XER-15 file Code Exicuted And \n Date is added And \n Sheets are added \n")     



def sr02():
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'SR-02'
    sr02_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'SR-08'
        if 'SR-02' in excel_file:
            sr02_file = excel_file
            print(f"File with 'SR-02' found: {sr02_file}")
            
# Check if a file with 'SR-02' was found
    if sr02_file:
        try:
        # Read the Excel file and get the value of cell A2
            sr02_file_path = os.path.join(output_folder_path, sr02_file)
            workbook = load_workbook(sr02_file_path)
            sheet = workbook.active
            sheet.title = 'SR02'
            print("Filename:", sr02_file)
            date = sr02_file[24:32]
            sheet = workbook.active
            print(f"Date: {date}")      # This will print Date from File name 
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
		
            # GREEN TABLE DATA FOR SR-02
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD DG"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Backward RD DG"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "N/A"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "N/A"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD DG",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Kohat Adda",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Kohat Adda",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD', 'Backward RD']
            L2 = ['FWD', 'BWD', 'FWD RD', 'BWD RD']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        except Exception as e:
            print(f"Error reading {sr02_file}: {e}")
    else:
        print("No file with 'SR-02' found. \n")
    workbook.save(sr02_file_path)    
    print("SR-02 file Code Exicuted And \n Date is added And \n Sheets are added \n")  





                                        #      /////////////////////////////////  ROUTES FOR MOH  ///////////////////////////////////////////////  

#/////////////////////////////////////////////////////////////////  ROUTES FOR MOH  /////////////////////////////////////////////////////////////////#


def sr08():
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'SR-08'
    sr08_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'SR-08'
        if 'SR-08' in excel_file:
            sr08_file = excel_file
            print(f"File with 'SR-08' found: {sr08_file}")
            
# Check if a file with 'SR-08' was found
    if sr08_file:
        try:
        # Read the Excel file and get the value of cell A2
            sr08_file_path = os.path.join(output_folder_path, sr08_file)
            workbook = load_workbook(sr08_file_path)
            sheet = workbook.active
            sheet.title = 'SR08'
            print("Filename:", sr08_file)
            date = sr08_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            sheet = workbook.active
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
		
            # GREEN TABLE DATA FOR SR-02
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "N/A"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "N/A"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD*",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Kohat Adda",$O:$O,"V*")`'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Kohat Adda",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD DG",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD', 'Backward RD']
            L2 = ['FWD', 'BWD', 'FWD RD', 'BWD RD']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        except Exception as e:
            print(f"Error reading {sr08_file}: {e}")
    else:
        print("No file with 'SR-02' found. \n")
    workbook.save(sr08_file_path)    
    print("SR-08 file Code Exicuted And \n Date is added And \n Sheets are added \n")  


def dr05():
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'SR-08'
    dr05_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'SR-08'
        if 'DR-05' in excel_file:
            dr05_file = excel_file
            print(f"File with 'DR-05' found: {dr05_file}")
            
# Check if a file with 'SR-08' was found
    if dr05_file:
        try:
        # Read the Excel file and get the value of cell A2
            dr05_file_path = os.path.join(output_folder_path, dr05_file)
            workbook = load_workbook(dr05_file_path)
            sheet = workbook.active
            sheet.title = 'DR05'
            date = dr05_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            sheet = workbook.active
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
		
            # GREEN TABLE DATA FOR ER-11
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD MOH"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD Phase 6"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD MOH"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD Phase 6"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD MOH",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Phase 6",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Phase 6",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD MOH', 'Backward RD MOH', 'Forward RD Phase 6', 'Backward RD Phase 6']
            L2 = ['FWD', 'BWD', 'FWD RD MOH', 'BWD RD MOH', 'FWD RD PH6', 'BWD RD PH6']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD MOH']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD MOH']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD PH6']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD PH6']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        except Exception as e:
            print(f"Error reading {dr05_file}: {e}")
    else:
        print("No file with 'DR-05' found. \n")
    workbook.save(dr05_file_path)    
    print("DR-05 file Code Exicuted And \n Date is added And \n Sheets are added \n") 


def dr06():
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'SR-08'
    sr08_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'SR-08'
        if 'DR-06' in excel_file:
            sr08_file = excel_file
            print(f"File with 'DR-06' found: {sr08_file}")
            
# Check if a file with 'SR-08' was found
    if sr08_file:
        try:
        # Read the Excel file and get the value of cell A2
            sr08_file_path = os.path.join(output_folder_path, sr08_file)
            workbook = load_workbook(sr08_file_path)
            sheet = workbook.active
            sheet.title = 'DR06'
            print("Filename:", sr08_file)
            date = sr08_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            sheet = workbook.active
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
		
  # GREEN TABLE DATA FOR ER-11
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD MOH"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD Phase 7"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD MOH"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD Phase 7"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD MOH",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Phase 7",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Phase 7",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD Phase 7",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD Phase 7",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD MOH', 'Backward RD MOH', 'Forward RD Phase 7', 'Backward RD Phase 7']
            L2 = ['FWD', 'BWD', 'FWD RD MOH', 'BWD RD MOH', 'FWD RD PH7', 'BWD RD PH7']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD MOH']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD MOH']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD PH7']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD PH7']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        except Exception as e:
            print(f"Error reading {sr08_file}: {e}")
    else:
        print("No file with 'DR-06' found. \n")
    workbook.save(sr08_file_path)    
    print("DR-06 file Code Exicuted And \n Date is added And \n Sheets are added \n")  


def dr07():
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'SR-08'
    sr08_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'SR-08'
        if 'DR-07' in excel_file:
            sr08_file = excel_file
            print(f"File with 'DR-07' found: {sr08_file}")
            
# Check if a file with 'SR-08' was found
    if sr08_file:
        try:
        # Read the Excel file and get the value of cell A2
            sr08_file_path = os.path.join(output_folder_path, sr08_file)
            workbook = load_workbook(sr08_file_path)
            sheet = workbook.active
            sheet.title = 'DR07'
            print("Filename:", sr08_file)
            date = sr08_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            sheet = workbook.active
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
		
            # GREEN TABLE DATA FOR ER-11
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD Karkhano"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD Phase 7"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD Karkhano"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD Phase 7"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Karkhano",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Karkhano",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Phase 7",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Phase 7",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD Phase 7",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD Phase 7",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD Karkhano', 'Backward RD Karkhano']
            L2 = ['FWD', 'BWD', 'FWD RD', 'BWD RD']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        except Exception as e:
            print(f"Error reading {sr08_file}: {e}")
    else:
        print("No file with 'DR-07' found. \n")
    workbook.save(sr08_file_path)    
    print("DR-07 file Code Exicuted And \n Date is added And \n Sheets are added \n") 


def dr11():

   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'SR-08'
    sr08_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'SR-08'
        if 'DR-11' in excel_file:
            sr08_file = excel_file
            print(f"File with 'DR-11' found: {sr08_file}")
            
# Check if a file with 'SR-08' was found
    if sr08_file:
        try:
        # Read the Excel file and get the value of cell A2
            sr08_file_path = os.path.join(output_folder_path, sr08_file)
            workbook = load_workbook(sr08_file_path)
            sheet = workbook.active
            sheet.title = 'DR11'
            print("Filename:", sr08_file)
            date = sr08_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            sheet = workbook.active
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
		
           # GREEN TABLE DATA FOR ER-11
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD MOH"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD Phase 6"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD MOH"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD Phase 6"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD MOH",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Phase 6",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Phase 6",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD MOH', 'Backward RD MOH', 'Forward RD Phase 6']
            L2 = ['FWD', 'BWD', 'FWD RD MOH', 'BWD RD MOH', 'FWD RD PH6']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD MOH']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD MOH']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD PH6']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
        except Exception as e:
            print(f"Error reading {sr08_file}: {e}")
    else:
        print("No file with 'DR-11' found. \n")
    workbook.save(sr08_file_path)    
    print("DR-11 file Code Exicuted And \n Date is added And \n Sheets are added \n") 


def er09():

    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'SR-08'
    sr08_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'SR-08'
        if 'ER-09' in excel_file:
            sr08_file = excel_file
            print(f"File with 'ER-09' found: {sr08_file}")
            
# Check if a file with 'SR-08' was found
    if sr08_file:
        try:
        # Read the Excel file and get the value of cell A2
            sr08_file_path = os.path.join(output_folder_path, sr08_file)
            workbook = load_workbook(sr08_file_path)
            sheet = workbook.active
            sheet.title = 'ER09'
            print("Filename:", sr08_file)
            date = sr08_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            sheet = workbook.active
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
		
            # GREEN TABLE DATA FOR ER-09
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD GLB"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD Phase 6"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD GLB"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD Phase 6"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD GLB",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD GLB",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Phase 6",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Phase 6",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD Phase 6",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD Phase 6",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD GLB', 'Backward RD GLB']
            L2 = ['FWD', 'BWD', 'FWD RD GLB', 'BWD RD GLB']
            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)
            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
                    # Select the "FWD" sheet
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD GLB']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD GLB']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

        except Exception as e:
            print(f"Error reading {sr08_file}: {e}")
    else:
        print("No file with 'ER-09' found. \n")
    workbook.save(sr08_file_path)    
    print("ER-09 file Code Exicuted And \n Date is added And \n Sheets are added \n")  


def er16():
   
   
    # List all files in the input folder
    files = [f for f in os.listdir(output_folder_path) if os.path.isfile(os.path.join(output_folder_path, f))]
# Filter Excel files
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xls")]
    
# Initialize the variable to store the file name containing 'SR-08'
    sr08_file = None
# Process each Excel file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder_path, excel_file)
        
    # Check if the file name contains 'SR-08'
        if 'ER-16' in excel_file:
            sr08_file = excel_file
            print(f"File with 'ER-16' found: {sr08_file}")
            
# Check if a file with 'SR-08' was found
    if sr08_file:
        try:
        # Read the Excel file and get the value of cell A2
            sr08_file_path = os.path.join(output_folder_path, sr08_file)
            workbook = load_workbook(sr08_file_path)
            sheet = workbook.active
            sheet.title = 'ER16'
            print("Filename:", sr08_file)
            date = sr08_file[24:32]
            print(f"Date: {date}")      # This will print Date from File name 
            sheet = workbook.active
            c = sheet.cell(row=1, column=1)
            # Inserting data to the cells 
            c.value = date  # This will add date to the 1st Cell
		
           # GREEN TABLE DATA FOR ER-16
            
            # CELL DATA FOR FOR 12M
            c8 =sheet.cell(row=8, column=21)
            c8.value = "Forward RD GC"
            c8 =sheet.cell(row=11, column=21)
            c8.value = "Forward RD Phase 6"
            
            c8 =sheet.cell(row=8, column=22)
            c8.value = "Backward RD GC"
            c8 =sheet.cell(row=11, column=22)
            c8.value = "Backward RD Phase 6"

            sheet['U4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['V4'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['U9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD GC",$O:$O,"V*")'
            sheet['V9'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD GC",$O:$O,"V*")'
            sheet['U12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Forward RD Phase 6",$O:$O,"V*")'
            sheet['V12'] = '=COUNTIFS($A:$A,"B*",$D:$D,"Backward RD Phase 6",$O:$O,"V*")'
            
            # CELL DATA FOR FOR 18M
            c8 =sheet.cell(row=8, column=23)
            c8.value = "Forward"
            c8 =sheet.cell(row=11, column=23)
            c8.value = "Forward"
            
            c8 =sheet.cell(row=8, column=24)
            c8.value = "Backward"
            c8 =sheet.cell(row=11, column=24)
            c8.value = "Backward"
            
            sheet['W4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward",$O:$O,"V*")'
            sheet['X4'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward",$O:$O,"V*")'
            sheet['W9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Forward RD",$O:$O,"V*")'
            sheet['X9'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD",$O:$O,"V*")'
            sheet['W12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            sheet['X12'] = '=COUNTIFS($A:$A,"A*",$D:$D,"Backward RD MOH",$O:$O,"V*")'
            
            #TOTAL
            sheet['Y4'] = '=SUM(U4:X4)'
            sheet['Y9'] = '=SUM(U9:X9)'
            sheet['Y12'] = '=SUM(U12:X12)'
            sheet['Y13'] = '=SUM(Y4:Y12)'
            # Read data into a pandas DataFrame, skipping the first row (header is in the 2nd row)
            data = sheet.iter_rows(values_only=True)
            next(data)  # Skip the first row
            columns = next(data)
            # Print the column names
            print(columns)
            # Convert the remaining rows into a DataFrame
            df = pd.DataFrame(data, columns=columns)
            L1 = ['Forward', 'Backward', 'Forward RD GC', 'Backward RD GC']
            L2 = ['FWD', 'BWD', 'FWD RD GC', 'BWD RD GC']
          # Define the range of columns to delete (R to Y)

            for driving_scheme, sheet_name in zip(L1, L2):
                # Filter rows based on the current driving_scheme
                filtered_df = df[df['Driving Scheme'].str.strip().eq(driving_scheme)]
                # Create a new sheet with the corresponding sheet_name and paste filtered data
                new_sheet = workbook.create_sheet(sheet_name)
                # Append column names
                new_sheet.append(df.columns.tolist())
                # Apply the formula to all cells in column H from H3 to the last filled row
                # Append rows from the filtered DataFrame
                for row in dataframe_to_rows(filtered_df, index=False, header=False):
                    new_sheet.append(row)

            print('\n Sheet making Loop is over now last loop should start exicuution \n')

            # Iterate over each sheet name in L2
            start_column = 'R'
            end_column = 'Y'
            for sheet_name in L2:
                # Get the corresponding sheet
                sheet = workbook[sheet_name]
                 # Insert a row at the top (before the existing rows)
                sheet.insert_rows(1)
                c = sheet.cell(row=1, column=1)
                c.value=date #this will add date to 1st Cell
                sheet['A1'].font = Font(b=True)
                    # Calculate the column indexes to delete
                start_index = sheet[start_column][0].column
                end_index = sheet[end_column][0].column + 1
    
                # Delete columns R to Y
                for _ in range(end_index - start_index):
                    sheet.delete_cols(start_index)
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            sheet = workbook['BWD']
            # Create a named style for the time format
            time_style = NamedStyle(name='time_format')
            time_style.number_format = '[hh]:mm'
            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['FWD RD GC']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'

            sheet = workbook['BWD RD GC']
            # Create a named style for the time format

            # Apply the named style to column H
            for cell in sheet['H']:
                cell.style = time_style
# Get the maximum row number
            max_row = sheet.max_row

            # Apply the formula to each row in column H from row 3 to the maximum row
            for row in range(3, max_row + 1):
                cell_g = sheet[f'G{row}']
                cell_f = sheet[f'F{row}']
                cell_h = sheet[f'H{row}']
                cell_h.value = f'={cell_g.coordinate}-{cell_f.coordinate}'
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
                
        except Exception as e:
            print(f"Error reading {sr08_file}: {e}")
    else:
        print("No file with 'ER-16' found. \n")
    workbook.save(sr08_file_path)    
    print("ER-16 file Code Exicuted And \n Date is added And \n Sheets are added \n")  

#

                                                                ##################################
                                                                ##################################
                                                                #     FUNCTION DEFINE SPACE      #
                                                                ##################################
                                                                ##################################
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Specifimg the input and output folder paths
input_folder_path = 'CONVERTED Waybills Chamkani xlsx'
#output_folder_path = "\\10.62.10.21\Users\Public\Master_Files\Reconciliation LMKR\Yaseen Ayaz\2024 files\Recon Genrated by Python Code"
output_folder_path = f"\\\\10.62.10.21\\Users\\Public\Master_Files\\Reconciliation LMKR\\Recon file Genrated By Py Code\\Recon for {date} {month_name} {year}"

# Creating the output folder if it doesn't exist
os.makedirs(output_folder_path, exist_ok=True)
#Calling the Function process_excel_files_in_folder() with two arguments input_folder_path, output_folder_path 
#The variables input_folder_path, output_folder_path data will be sended to the function for further working.  
process_excel_files_in_folder(input_folder_path, output_folder_path)
#/////////////////////////////////////////////////////////
# Use shutil.rmtree to delete the directory CONVERTED Waybills Chamkani xlsx and its contents


# Use os.listdir to get a list of files in the output folder
files_in_output_folder = os.listdir(output_folder_path)
# Create an empty list to store file names
file_names_list = []
# Iterate over the files and print the names
print("\n File Names in the Folder Are: \n")
for file_name in files_in_output_folder:
    print(file_name)
    # Append each file name to the list
    file_names_list.append(file_name)

print('\n') 

#/////////////////////////////////////////////////////////////////
#/////////////////////////////////////////////////////////////////
# Check for every possible combination of file names for MOH ROUTES
# Check if "DR-05" is present in any part of the list elements
target_string = "DR-05"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    dr05()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "DR-06" is present in any part of the list elements
target_string = "DR-06"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    dr06()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "DR-07" is present in any part of the list elements
target_string = "DR-07"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    dr07()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "ER-09" is present in any part of the list elements
target_string = "ER-09"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    er09()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "ER-16" is present in any part of the list elements
target_string = "ER-16"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    er16()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "SR-08" is present in any part of the list elements
target_string = "DR-11"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    dr11()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "SR-08" is present in any part of the list elements
target_string = "SR-08"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    sr08()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
#/////////////////////////////////////////////////////////////////
#/////////////////////////////////////////////////////////////////
# Check for every possible combination of file names for CHAMKANI ROUTES
print('\n')
# Check if "SR-02" is present in any part of the list elements
target_string = "SR-02"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    sr02()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")

# Check if "DR-03A" is present in any part of the list elements
target_string = "DR-03A"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    dr3a()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "DR-03B" is present in any part of the list elements
target_string = "DR-03B"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    dr3b()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "DR-04B" is present in any part of the list elements
target_string = "DR-04B"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    dr4b()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "ER-01" is present in any part of the list elements
target_string = "ER-01"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    er01()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "ER-10" is present in any part of the list elements
target_string = "ER-10"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    er10()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "ER-12" is present in any part of the list elements
target_string = "ER-12"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    er12()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")
# Check if "XER-15" is present in any part of the list elements
target_string = "XER-15"
found = any(target_string in file_name for file_name in file_names_list)
if found:
    print(f"Yes, '{target_string}' is present in at least one file name.")
    xer15()
else:
    print(f"No, '{target_string}' is not present in any file name. \n")

shutil.rmtree(output_folder)

#////////////////////////////////////////////////////////////////////////////
# It will make a rar file from final output folder.
def zip_folder(folder_path, zip_file_name):
    # Check if the folder exists
    if not os.path.exists(folder_path):
        print(f"Folder '{folder_path}' does not exist.")
        return
    
    # Create a ZipFile object in write mode
    with zipfile.ZipFile(zip_file_name, 'w') as zipf:
        # Iterate over all the files in the folder
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            # Add the file to the zip file with its relative path
            zipf.write(file_path, os.path.basename(file_path))

    print(f"Folder '{folder_path}' successfully zipped to '{zip_file_name}'.")

# Provide the path to the folder you want to zip
folder_to_zip = output_folder_path

# Provide the name for the resulting zip file
zip_file_name = f"\\\\10.62.10.21\\Users\\Public\Master_Files\\Reconciliation LMKR\\Recon file Genrated By Py Code\\Recon for {date} {month_name} {year}.zip"


# Call the function to zip the folder
zip_folder(folder_to_zip, zip_file_name)
